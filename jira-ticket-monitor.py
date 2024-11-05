import requests
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter  # Add this import
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import configparser

# Load configuration from ini file
config = configparser.ConfigParser()
config.read('config.ini')

# Jira API URL and authentication details
JIRA_URL = config['JIRA']['URL']
AUTH = HTTPBasicAuth(config['JIRA']['USERNAME'], config['JIRA']['API_TOKEN'])
HEADERS = {"Accept": "application/json"}

# Jira Query Language (JQL) query to fetch issues
JQL_QUERY = config['JIRA']['JQL_QUERY']
MAX_RESULTS = int(config['JIRA']['MAX_RESULTS'])

# Number of days to highlight recent comments
HIGHLIGHT_DAYS = int(config['SETTINGS']['HIGHLIGHT_DAYS'])

# Number of recent comments to include in the Excel report
RECENT_COMMENTS_COUNT = int(config['SETTINGS']['RECENT_COMMENTS_COUNT'])

# File name prefix for the Excel report
FILE_NAME_PREFIX = config['SETTINGS']['FILE_NAME_PREFIX']

def fetch_issues():
    """Fetch all issues from Jira based on the JQL query."""
    start_at = 0
    all_issues = []
    while True:
        params = {'jql': JQL_QUERY, 'maxResults': MAX_RESULTS, 'startAt': start_at}
        response = requests.get(JIRA_URL, headers=HEADERS, auth=AUTH, params=params)
        response.raise_for_status()
        issues = response.json().get('issues', [])
        if not issues:
            break
        all_issues.extend(issues)
        start_at += MAX_RESULTS
    return all_issues

def fetch_comments(issue_key):
    """Fetch the last few comments for a given Jira issue."""
    comments_url = f"https://metainfra.atlassian.net/rest/api/3/issue/{issue_key}"
    response = requests.get(comments_url, headers=HEADERS, auth=AUTH)
    response.raise_for_status()
    comments_data = response.json().get('fields', {}).get('comment', {}).get('comments', [])
    comments_list = []

    def extract_text(content):
        """Recursively extract text from the content."""
        if isinstance(content, list):
            return ''.join(extract_text(item) for item in content)

        if not isinstance(content, dict):
            return str(content)

        if content['type'] == 'text':
            return content['text']
        elif content['type'] == 'mention':
            return content['attrs']['text']
        elif content['type'] == 'hardBreak':
            return "\n"
        elif 'content' in content:
            return ''.join(extract_text(item) for item in content['content'])
        return ""

    # Only take the last few comments based on RECENT_COMMENTS_COUNT
    for comment in comments_data[-RECENT_COMMENTS_COUNT:][::-1]:
        try:
            author = comment['updateAuthor']['displayName']
            update_time = comment['updated']
            comment_body = extract_text(comment['body']['content'])

            # Combine the full comment content
            # Convert update time to local timezone
            update_time = datetime.strptime(update_time, '%Y-%m-%dT%H:%M:%S.%f%z')
            local_update_time = update_time.astimezone().strftime('%Y-%m-%d %H:%M:%S')
            full_comment = f"**{author} ({local_update_time}):**\n{comment_body}"
            comments_list.append(full_comment)
        except (KeyError, IndexError) as e:
            comments_list.append(f"Error parsing comment: {str(e)}")

    return comments_list  # Return the list of comments, not a combined string

def extract_labels(issue, prefix):
    """Extract labels from an issue based on a given prefix."""
    labels = [label[len(prefix):] for label in issue['fields']['labels'] if label.startswith(prefix)]
    return ','.join(labels)

def create_excel(issues):
    """Create an Excel file with the fetched Jira issues and their details."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jira Issues"
    
    # Insert the JQL query from config.ini into the first row
    ws.append([f"JQL Query: {JQL_QUERY}"])
    
    # Merge the cells for the JQL query row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)  # Adjusted end_column to 9
    
    # Add the header row for the issues
    ws.append(["Jira Ticket ID", "Summary", "PIC", "Status", "Priority", "Update Time", "Sensor Issue Category", "Gerrit ID", "Comments"])
    wrap_alignment = Alignment(wrap_text=True)

    # Apply the same formatting to the first row as the second row
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.font = Font(bold=True, color=cell.font.color)
        cell.alignment = wrap_alignment

    with ThreadPoolExecutor(max_workers=20) as executor:
        future_to_issue = {executor.submit(fetch_comments, issue['key']): issue for issue in issues}
        for index, future in enumerate(as_completed(future_to_issue), start=1):
            issue = future_to_issue[future]
            issue_key = issue['key']
            summary = issue['fields']['summary']
            assignee = issue['fields']['assignee']['displayName'] if issue['fields']['assignee'] else 'Unassigned'
            status = issue['fields']['status']['name']
            priority = issue['fields']['priority']['name'] if issue['fields']['priority'] else 'None'
            updated = issue['fields']['updated']
            comments = future.result()  # 這裡獲取的是評論列表

            # Extract labels
            sensor_issue_category = extract_labels(issue, 'issue-category:')
            gerrit_id = extract_labels(issue, 'gerrit:')

            # Convert update time to local timezone
            update_time = datetime.strptime(updated, '%Y-%m-%dT%H:%M:%S.%f%z')
            local_update_time = update_time.astimezone().strftime('%Y-%m-%d %H:%M:%S')

            # Debug message
            # print(f"Processing issue {index}/{len(issues)}: {issue_key}")

            # 第一行
            start_row = ws.max_row + 1
            ws.append([issue_key, summary, assignee, status, priority, local_update_time, sensor_issue_category, gerrit_id, comments[0] if comments else ""])

            # 設置超連結
            cell = ws.cell(row=start_row, column=1)
            cell.value = issue_key
            cell.hyperlink = f"https://metainfra.atlassian.net/browse/{issue_key}"
            cell.style = "Hyperlink"

            # 添加其餘評論
            for comment in comments[1:]:
                ws.append([issue_key, summary, assignee, status, priority, local_update_time, sensor_issue_category, gerrit_id, comment])

            # 合併相同 issue 的儲存格
            end_row = ws.max_row
            # for col in range(1, 9):  # 合併 A 到 H 欄
            #     if start_row != end_row:  # 只在有多行時才合併
            #         ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

            # 設置評論的顏色
            for row in range(start_row, end_row + 1):
                comment_cell = ws.cell(row=row, column=9)
                if comment_cell.value and '(' in comment_cell.value and ')' in comment_cell.value:
                    comment_time_str = comment_cell.value.split('(')[1].split(')')[0]
                    try:
                        comment_time = datetime.strptime(comment_time_str, '%Y-%m-%d %H:%M:%S')
                        if (datetime.now(comment_time.tzinfo) - comment_time).days <= HIGHLIGHT_DAYS:
                            comment_cell.font = Font(color="0000FF")
                    except (ValueError, IndexError):
                        pass

            progress = (index / len(issues)) * 100
            print(f"Processed {index}/{len(issues)} issues ({progress:.2f}%)")

    format_excel(ws)
    save_excel(wb)

def format_excel(ws):
    """Format the Excel sheet with appropriate styles and widths."""
    
    # 設置單元格對齊方式為自動換行和頂部對齊
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap_alignment
            cell.font = Font(name='Calibri', color=cell.font.color, bold=cell.font.bold, italic=cell.font.italic, underline=cell.font.underline)

    # 設置第一行和第二行的背景顏色和字體加粗
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=ws.max_column):  # 遍歷第一行和第二行
        for cell in row:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(bold=True, color=cell.font.color)

    # 調整每一欄的寬度
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        adjusted_width = (max_length + 2)
        col_letter = get_column_letter(col[0].column)  # 使用 get_column_letter
        ws.column_dimensions[col_letter].width = adjusted_width

    # 設置特定欄的寬度
    ws.column_dimensions['A'].width = 14.86
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['G'].width = 30  # Adjusted width for new columns
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 100

    # 設置單元格邊框
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # 對第一列（A列）應用超連結樣式，除了標題行
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(color="0000FF", underline="single", name=cell.font.name, bold=cell.font.bold, italic=cell.font.italic)

def save_excel(wb):
    """Save the Excel workbook to a file."""
    current_time = datetime.now().strftime("%m-%d_%H.%M")
    file_name = f"{FILE_NAME_PREFIX}_jira_issues_{current_time}.xlsx"
    file_path = rf"C:\Users\Wes\Documents\PlatformIO\Projects\jira-ticket-monitor\{file_name}"
    wb.save(file_path)
    print(f"Jira issues have been written to {file_name}")

def main():
    """Main function to fetch Jira issues and create an Excel report."""
    print("Sending request to Jira...")
    try:
        issues = fetch_issues()
        print(f"Total issues to process: {len(issues)}")
        create_excel(issues)
    except requests.RequestException as e:
        print(f"Failed to fetch issues: {e}")

if __name__ == "__main__":
    main()
